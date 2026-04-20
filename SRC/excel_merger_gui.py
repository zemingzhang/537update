"""
带颜色高亮的Excel文件合并工具（图形界面版）
功能：比较两个Excel文件的第一列（商品名称），将文件2中独有的行添加到文件1，合并重复项的订单量
新增功能：合并结果中新增项用黄色标记，合并项用绿色标记
依赖：pandas, openpyxl, tkinter（均为标准库或已安装库）
使用方法：双击运行此文件
"""

import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import threading
import urllib.request
import urllib.error
import urllib.parse
import json
import sys
import subprocess

# ===== 版本与更新配置 =====
# 当前版本号，需与 GitHub 上 version.json 中的 version 字段格式一致
CURRENT_VERSION = "v2.0"
GITHUB_REPO = "zemingzhang/537update"
# 通过 GitHub API 自动获取默认分支，再读取 version.json（无需手动指定 main/master）
GITHUB_API_REPO_URL = f"https://api.github.com/repos/{GITHUB_REPO}"

class ExcelMergerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel文件合并工具 v2.0 (带颜色高亮)")
        self.root.geometry("780x720")
        
        # 设置窗口图标（可选）
        try:
            self.root.iconbitmap(default='')
        except:
            pass
        
        # 初始化文件路径变量
        self.file1_path = ""
        self.file2_path = ""
        
        # 颜色定义
        self.new_item_color = PatternFill(start_color="FFFF00",  # 黄色
                                          end_color="FFFF00", 
                                          fill_type="solid")
        self.merged_item_color = PatternFill(start_color="90EE90",  # 浅绿色
                                             end_color="90EE90", 
                                             fill_type="solid")
        
        # 创建界面组件
        self.setup_ui()
    
    def setup_ui(self):
        """创建和布局所有界面组件"""
        # ===== 顶部更新栏 =====
        top_bar = tk.Frame(self.root, bg="#e8e8e8", bd=1, relief=tk.GROOVE)
        top_bar.pack(fill=tk.X, padx=0, pady=(0, 5))

        self.update_version_label = tk.Label(
            top_bar,
            text=f"当前版本: {CURRENT_VERSION}",
            font=("微软雅黑", 8),
            fg="gray",
            bg="#e8e8e8"
        )
        self.update_version_label.pack(side=tk.LEFT, padx=10, pady=4)

        self.update_btn = ttk.Button(
            top_bar,
            text="🔄 检查更新",
            command=self.check_for_updates,
            width=12
        )
        self.update_btn.pack(side=tk.RIGHT, padx=10, pady=3)

        # 标题区域
        title_frame = tk.Frame(self.root)
        title_frame.pack(pady=20, fill=tk.X)
        
        title_label = tk.Label(title_frame, 
                               text="🎨 Excel文件合并工具 (带颜色高亮)", 
                               font=("微软雅黑", 18, "bold"))
        title_label.pack()
        
        info_label = tk.Label(title_frame,
                              text="请选择两个需要合并的Excel文件（需包含'商品'、'盒码'、'订单量'三列）",
                              font=("微软雅黑", 10),
                              wraplength=500,
                              fg="gray")
        info_label.pack(pady=5)
        
        # 颜色说明区域
        color_frame = tk.Frame(self.root)
        color_frame.pack(pady=5, padx=20, fill=tk.X)
        
        color_label = tk.Label(color_frame,
                              text="颜色说明：",
                              font=("微软雅黑", 10, "bold"))
        color_label.pack(anchor=tk.W)
        
        # 新增项颜色说明
        new_color_frame = tk.Frame(color_frame)
        new_color_frame.pack(anchor=tk.W, pady=2)
        
        new_color_box = tk.Label(new_color_frame, 
                                text="  ", 
                                bg="yellow", 
                                width=3, 
                                height=1)
        new_color_box.pack(side=tk.LEFT, padx=(0, 5))
        
        new_color_text = tk.Label(new_color_frame, 
                                 text="新增项（文件2中有，文件1中没有的商品）", 
                                 font=("微软雅黑", 9))
        new_color_text.pack(side=tk.LEFT)
        
        # 合并项颜色说明
        merged_color_frame = tk.Frame(color_frame)
        merged_color_frame.pack(anchor=tk.W, pady=2)
        
        merged_color_box = tk.Label(merged_color_frame, 
                                   text="  ", 
                                   bg="#90EE90", 
                                   width=3, 
                                   height=1)
        merged_color_box.pack(side=tk.LEFT, padx=(0, 5))
        
        merged_color_text = tk.Label(merged_color_frame, 
                                    text="合并项（两个文件中都存在的商品，订单量已相加）", 
                                    font=("微软雅黑", 9))
        merged_color_text.pack(side=tk.LEFT)
        
        # 文件选择区域
        self.create_file_selection("第一个Excel文件", 0)
        self.create_file_selection("第二个Excel文件", 1)
        
        # 状态显示区域
        status_frame = tk.LabelFrame(self.root, text="处理状态", font=("微软雅黑", 10))
        status_frame.pack(pady=15, padx=20, fill=tk.BOTH, expand=True)
        
        # 添加滚动条
        scrollbar = tk.Scrollbar(status_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.status_text = tk.Text(status_frame, 
                                  height=8, 
                                  width=70,
                                  font=("宋体", 9),
                                  yscrollcommand=scrollbar.set)
        self.status_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5, pady=5)
        scrollbar.config(command=self.status_text.yview)
        
        # 初始状态信息
        self.update_status("就绪。请选择两个Excel文件。\n")
        self.status_text.config(state=tk.DISABLED)
        
        # 按钮区域
        button_frame = tk.Frame(self.root)
        button_frame.pack(pady=10)
        
        # 样式化合并按钮
        style = ttk.Style()
        style.configure("Large.TButton", font=("微软雅黑", 12), padding=10)
        
        self.merge_button = ttk.Button(button_frame, 
                                       text="开始合并文件", 
                                       command=self.merge_files,
                                       state=tk.DISABLED)
        self.merge_button.pack(side=tk.LEFT, padx=5)
        
        exit_button = ttk.Button(button_frame,
                                 text="退出程序",
                                 command=self.root.quit)
        exit_button.pack(side=tk.LEFT, padx=5)
        
        # 底部状态栏
        self.status_bar = tk.Label(self.root, 
                                   text="等待操作...", 
                                   font=("微软雅黑", 9),
                                   fg="gray",
                                   bd=1, 
                                   relief=tk.SUNKEN, 
                                   anchor=tk.W)
        self.status_bar.pack(side=tk.BOTTOM, fill=tk.X)
        
        # 绑定关闭事件
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
    
    def create_file_selection(self, label_text, file_index):
        """创建文件选择组件"""
        frame = tk.LabelFrame(self.root, text=f"选择{label_text}", font=("微软雅黑", 10))
        frame.pack(pady=10, padx=20, fill=tk.X)
        
        # 按钮和标签框架
        inner_frame = tk.Frame(frame)
        inner_frame.pack(pady=10, padx=10, fill=tk.X)
        
        # 选择按钮
        if file_index == 0:
            command = self.select_file1
            var_attr = "btn_file1"
            label_attr = "label_file1"
        else:
            command = self.select_file2
            var_attr = "btn_file2"
            label_attr = "label_file2"
        
        button = ttk.Button(inner_frame, 
                           text=f"浏览...", 
                           command=command,
                           width=15)
        button.pack(side=tk.LEFT, padx=(0, 10))
        setattr(self, var_attr, button)
        
        # 文件路径标签
        label = tk.Label(inner_frame, 
                        text="未选择文件", 
                        font=("微软雅黑", 9),
                        fg="gray",
                        bg="#f0f0f0",
                        relief=tk.SUNKEN,
                        anchor=tk.W,
                        width=50)
        label.pack(side=tk.LEFT, fill=tk.X, expand=True)
        setattr(self, label_attr, label)
    
    def select_file1(self):
        """选择第一个Excel文件"""
        filename = filedialog.askopenfilename(
            title="选择第一个Excel文件",
            filetypes=[("Excel文件", "*.xlsx *.xls"), ("所有文件", "*.*")]
        )
        if filename:
            self.file1_path = filename
            self.label_file1.config(text=os.path.basename(filename), fg="black")
            self.update_status(f"已选择文件1: {filename}")
            self.check_files_selected()
    
    def select_file2(self):
        """选择第二个Excel文件"""
        filename = filedialog.askopenfilename(
            title="选择第二个Excel文件",
            filetypes=[("Excel文件", "*.xlsx *.xls"), ("所有文件", "*.*")]
        )
        if filename:
            self.file2_path = filename
            self.label_file2.config(text=os.path.basename(filename), fg="black")
            self.update_status(f"已选择文件2: {filename}")
            self.check_files_selected()
    
    def check_files_selected(self):
        """检查是否已选择两个文件，如果是则启用合并按钮"""
        if self.file1_path and self.file2_path:
            self.merge_button.config(state=tk.NORMAL)
            self.status_bar.config(text="已选择两个文件，可以开始合并")
    
    def update_status(self, message):
        """在状态文本框添加新消息"""
        self.status_text.config(state=tk.NORMAL)
        self.status_text.insert(tk.END, f"> {message}\n")
        self.status_text.see(tk.END)  # 滚动到底部
        self.status_text.config(state=tk.DISABLED)
        self.root.update()  # 更新界面显示
    
    def apply_colors_to_excel(self, file_path, new_indices, merged_indices):
        """为Excel文件中的特定行应用颜色"""
        try:
            wb = load_workbook(file_path)
            ws = wb.active
            
            # 为新增项应用黄色
            for idx in new_indices:
                excel_row = idx + 2  # Excel行索引从1开始，表头占第1行
                for col in range(1, 4):  # 前三列
                    ws.cell(row=excel_row, column=col).fill = self.new_item_color
            
            # 为合并项应用绿色
            for idx in merged_indices:
                excel_row = idx + 2
                for col in range(1, 4):  # 前三列
                    ws.cell(row=excel_row, column=col).fill = self.merged_item_color
            
            wb.save(file_path)
            self.update_status("✅ 已为合并结果应用颜色标记")
            
        except Exception as e:
            self.update_status(f"⚠️ 应用颜色时出错: {str(e)}")
    
    def merge_files(self):
        """执行合并操作的核心函数"""
        if not (self.file1_path and self.file2_path):
            messagebox.showerror("错误", "请先选择两个Excel文件！")
            return
        
        # 禁用按钮，防止重复点击
        self.merge_button.config(state=tk.DISABLED, text="处理中...")
        self.status_bar.config(text="正在处理，请稍候...")
        self.root.update()
        
        try:
            # 读取文件
            self.update_status("正在读取文件...")
            df1 = pd.read_excel(self.file1_path)
            df2 = pd.read_excel(self.file2_path)
            
            # 检查必要列
            required_columns = ['商品', '盒码', '订单量']
            for df, name in [(df1, '文件1'), (df2, '文件2')]:
                missing = [col for col in required_columns if col not in df.columns]
                if missing:
                    raise ValueError(f"{name}缺少必要列: {missing}")
            
            self.update_status(f"文件1读取成功，共 {len(df1)} 行")
            self.update_status(f"文件2读取成功，共 {len(df2)} 行")
            
            # 处理逻辑
            result_df = df1.copy()
            df1_products = set(df1['商品'].astype(str).str.strip())
            df2_products = set(df2['商品'].astype(str).str.strip())
            
            # 找出新增商品
            new_products = df2_products - df1_products
            new_count = len(new_products)
            self.update_status(f"发现 {new_count} 个新增商品")
            
            # 记录新增项的索引
            new_indices = []
            if new_products:
                new_rows = df2[df2['商品'].astype(str).str.strip().isin(new_products)].copy()
                # 新增项的索引是原结果行数到原结果行数+新增行数-1
                start_idx = len(result_df)
                new_indices = list(range(start_idx, start_idx + len(new_rows)))
                result_df = pd.concat([result_df, new_rows], ignore_index=True, sort=False)
            
            # 处理重复商品，合并订单量
            common_products = df1_products.intersection(df2_products)
            common_count = len(common_products)
            self.update_status(f"发现 {common_count} 个共同商品，正在合并订单量...")
            
            # 创建商品到订单量的映射
            product_to_quantity = {}
            for _, row in df2.iterrows():
                product_name = str(row['商品']).strip()
                if product_name in common_products:
                    if product_name not in product_to_quantity:
                        product_to_quantity[product_name] = 0
                    product_to_quantity[product_name] += float(row['订单量'])
            
            # 更新结果中的订单量，并记录合并项的索引
            merged_indices = []
            merge_count = 0
            for idx, row in result_df.iterrows():
                product_name = str(row['商品']).strip()
                if product_name in product_to_quantity:
                    result_df.at[idx, '订单量'] = float(row['订单量']) + product_to_quantity[product_name]
                    merged_indices.append(idx)
                    merge_count += 1
            
            # 生成输出文件名（带时间戳）
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"合并结果_带颜色_{timestamp}.xlsx"
            
            # 让用户选择保存位置
            output_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                initialfile=output_filename,
                filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")]
            )
            
            if not output_path:  # 用户取消了保存
                self.update_status("操作已取消")
                self.merge_button.config(state=tk.NORMAL, text="开始合并文件")
                self.status_bar.config(text="操作取消")
                return
            
            # 保存结果
            result_df[['商品', '盒码', '订单量']].to_excel(output_path, index=False)
            
            # 为Excel文件添加颜色
            self.update_status("正在为合并结果添加颜色标记...")
            self.apply_colors_to_excel(output_path, new_indices, merged_indices)
            
            self.update_status("✅ 处理完成！")
            self.update_status(f"   总记录数: {len(result_df)} 行")
            self.update_status(f"   - 新增项（黄色）: {new_count} 行")
            self.update_status(f"   - 合并项（绿色）: {merge_count} 行")
            self.update_status(f"   文件已保存至: {output_path}")
            
            # 恢复按钮状态
            self.merge_button.config(state=tk.NORMAL, text="开始合并文件")
            self.status_bar.config(text=f"处理完成！结果保存为: {os.path.basename(output_path)}")
            
            # 询问是否打开文件
            if messagebox.askyesno("完成", 
                                  f"合并完成！\n\n"
                                  f"总记录数: {len(result_df)} 行\n"
                                  f"新增项（黄色）: {new_count} 行\n"
                                  f"合并项（绿色）: {merge_count} 行\n\n"
                                  f"是否要打开结果文件？"):
                try:
                    os.startfile(output_path)
                except:
                    messagebox.showinfo("打开文件", f"文件已保存到:\n{output_path}")
            
        except FileNotFoundError as e:
            error_msg = f"文件未找到: {str(e)}\n请检查文件路径是否正确。"
            self.update_status(f"❌ 错误: {error_msg}")
            messagebox.showerror("文件错误", error_msg)
        except PermissionError as e:
            error_msg = f"文件被占用或无权限: {str(e)}\n请关闭已打开的Excel文件。"
            self.update_status(f"❌ 错误: {error_msg}")
            messagebox.showerror("权限错误", error_msg)
        except ValueError as e:
            error_msg = f"文件格式错误: {str(e)}\n请确保Excel文件包含'商品'、'盒码'、'订单量'三列。"
            self.update_status(f"❌ 错误: {error_msg}")
            messagebox.showerror("格式错误", error_msg)
        except Exception as e:
            error_msg = f"处理过程中发生未知错误:\n{str(e)}"
            self.update_status(f"❌ 错误: {error_msg}")
            messagebox.showerror("处理错误", error_msg)
            
            # 恢复按钮状态
            self.merge_button.config(state=tk.NORMAL, text="开始合并文件")
            self.status_bar.config(text="处理出错，请检查文件格式")
    
    def on_closing(self):
        """关闭窗口时的确认"""
        if messagebox.askokcancel("退出", "确定要退出程序吗？"):
            self.root.destroy()

    # ===== 自动更新相关方法 =====
    # 设计说明：
    #   运行中的 EXE 文件被 Windows 锁定，无法直接自我替换，也无需额外的 update.exe。
    #   方案：主程序从 GitHub 下载新 EXE 到系统临时目录，然后生成一个临时批处理脚本，
    #         该脚本等待主进程 PID 退出后，将新 EXE 覆盖旧 EXE，并重新启动。

    def check_for_updates(self):
        """点击检查更新按钮：在后台线程中获取并解析 version.json"""
        self.update_btn.config(state=tk.DISABLED, text="检查中...")
        self.update_version_label.config(
            text=f"当前版本: {CURRENT_VERSION}  正在检查更新..."
        )
        threading.Thread(target=self._check_update_thread, daemon=True).start()

    def _check_update_thread(self):
        """后台线程：自动检测默认分支，再获取 version.json 比对版本号"""
        try:
            headers = {"User-Agent": f"ExcelMergerApp/{CURRENT_VERSION}"}

            # 第一步：通过 GitHub API 获取仓库默认分支名（main / master / 其他）
            repo_req = urllib.request.Request(GITHUB_API_REPO_URL, headers=headers)
            with urllib.request.urlopen(repo_req, timeout=10) as resp:
                repo_info = json.loads(resp.read().decode("utf-8"))
            default_branch = repo_info.get("default_branch", "main")

            # 第二步：用真实的分支名拼接 raw URL 读取 version.json
            version_url = (
                f"https://raw.githubusercontent.com/{GITHUB_REPO}"
                f"/{default_branch}/version.json"
            )
            ver_req = urllib.request.Request(version_url, headers=headers)
            with urllib.request.urlopen(ver_req, timeout=10) as resp:
                data = json.loads(resp.read().decode("utf-8"))

            latest_version = str(data.get("version", "")).strip()
            exe_url = str(data.get("exe_url", "")).strip()
            release_notes = str(data.get("release_notes", "")).strip()

            if not latest_version:
                raise ValueError("version.json 中缺少 version 字段")

            if latest_version != CURRENT_VERSION:
                self.root.after(0, lambda: self._on_update_available(
                    latest_version, exe_url, release_notes))
            else:
                self.root.after(0, self._on_no_update)

        except urllib.error.URLError as e:
            msg = str(e)
            self.root.after(0, lambda: self._on_update_error(f"网络连接失败: {msg}"))
        except Exception as e:
            msg = str(e)
            self.root.after(0, lambda: self._on_update_error(msg))

    def _on_update_available(self, new_version, exe_url, release_notes):
        """发现新版本：弹窗提示用户是否下载"""
        self.update_btn.config(state=tk.NORMAL, text="🔄 检查更新")
        self.update_version_label.config(
            text=f"当前版本: {CURRENT_VERSION}  ⬆ 发现新版本: {new_version}"
        )
        notes_section = f"\n\n更新内容:\n{release_notes}" if release_notes else ""
        msg = (f"发现新版本 {new_version}！\n"
               f"当前版本: {CURRENT_VERSION}"
               f"{notes_section}\n\n"
               f"是否立即下载并安装更新？\n"
               f"（程序将自动关闭，安装完成后自动重启）")
        if messagebox.askyesno("发现新版本", msg):
            if exe_url:
                self._start_download(exe_url, new_version)
            else:
                messagebox.showinfo(
                    "更新提示",
                    f"version.json 中未提供 exe_url，请前往 GitHub 手动下载：\n"
                    f"https://github.com/{GITHUB_REPO}/releases"
                )

    def _on_no_update(self):
        """已是最新版本"""
        self.update_btn.config(state=tk.NORMAL, text="🔄 检查更新")
        self.update_version_label.config(
            text=f"当前版本: {CURRENT_VERSION}  ✅ 已是最新版本"
        )
        messagebox.showinfo("检查更新", "当前已是最新版本！")

    def _on_update_error(self, error_msg):
        """检查或下载出错时恢复界面并提示"""
        self.update_btn.config(state=tk.NORMAL, text="🔄 检查更新")
        self.update_version_label.config(
            text=f"当前版本: {CURRENT_VERSION}  ⚠ 检查失败"
        )
        messagebox.showerror("检查更新失败", f"无法检查更新:\n{error_msg}")

    def _start_download(self, exe_url, new_version):
        """启动后台下载线程"""
        self.update_btn.config(state=tk.DISABLED, text="下载中...")
        self.update_version_label.config(
            text=f"当前版本: {CURRENT_VERSION}  正在下载 {new_version} 0%..."
        )
        threading.Thread(
            target=self._download_thread,
            args=(exe_url, new_version),
            daemon=True
        ).start()

    def _http_get_follow_redirects(self, url, headers, timeout=120):
        """
        使用自定义重定向处理器下载文件，确保每次重定向都保留完整请求头。
        urllib 默认在重定向时会丢弃自定义头（如 Accept），导致 GitHub CDN
        返回 HTML 而非 EXE，此方法通过覆盖 redirect_request 解决该问题。
        """
        original_headers = dict(headers)

        class _KeepHeadersRedirectHandler(urllib.request.HTTPRedirectHandler):
            def redirect_request(self, req, fp, code, msg, resp_headers, newurl):
                new_req = super().redirect_request(
                    req, fp, code, msg, resp_headers, newurl)
                if new_req is not None:
                    for key, val in original_headers.items():
                        new_req.add_unredirected_header(key, val)
                return new_req

        opener = urllib.request.build_opener(_KeepHeadersRedirectHandler())
        req = urllib.request.Request(url, headers=headers)
        resp = opener.open(req, timeout=timeout)
        return resp, resp.url

    def _download_thread(self, exe_url, new_version):
        """后台线程：分块下载新 EXE，然后生成替换脚本"""
        import tempfile

        # 兼容 PyInstaller 打包的 EXE 和直接运行的 .py
        if getattr(sys, "frozen", False):
            current_exe = sys.executable  # 打包后为当前 .exe 的完整路径
        else:
            current_exe = os.path.abspath(__file__)

        tmp_dir = tempfile.gettempdir()
        new_exe_path = os.path.join(tmp_dir, "excel_merger_new.exe")
        bat_path = os.path.join(tmp_dir, "excel_merger_update.bat")

        try:
            headers = {
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
                "Accept": "application/octet-stream, */*",
            }
            # 手动跟踪重定向，保证每一跳都发送完整请求头
            resp, final_url = self._http_get_follow_redirects(exe_url, headers)

            content_type = resp.headers.get("Content-Type", "")
            if "text/html" in content_type:
                raise ValueError(
                    f"下载到的是 HTML 页面而非 EXE\n"
                    f"最终请求地址: {final_url}\n"
                    f"请确认 Release 为公开仓库且 exe_url 指向文件本身"
                )

            total = int(resp.headers.get("Content-Length", 0) or 0)
            downloaded = 0
            chunk_size = 65536  # 64 KB
            with open(new_exe_path, "wb") as f:
                while True:
                    chunk = resp.read(chunk_size)
                    if not chunk:
                        break
                    f.write(chunk)
                    downloaded += len(chunk)
                    if total > 0:
                        pct = int(downloaded * 100 / total)
                        self.root.after(0, lambda p=pct: self.update_version_label.config(
                            text=f"当前版本: {CURRENT_VERSION}  下载中 {p}%..."
                        ))

            # 校验：有效的 Windows PE 可执行文件以 MZ 开头
            with open(new_exe_path, "rb") as f:
                magic = f.read(2)
            if magic != b"MZ":
                raise ValueError(
                    f"下载的文件不是有效的 EXE（文件头: {magic.hex()}，非 MZ）\n"
                    f"文件大小: {downloaded} 字节，最终地址: {final_url}"
                )

            # 生成等待主进程退出后执行替换并重启的批处理脚本
            # 注意：路径中若含空格，move 命令已用双引号包裹
            pid = os.getpid()
            bat_lines = [
                "@echo off",
                ":waitloop",
                f'tasklist /fi "PID eq {pid}" 2>nul | findstr /I "{pid}" >nul 2>&1',
                "if not errorlevel 1 (",
                "    timeout /t 1 /nobreak >nul",
                "    goto waitloop",
                ")",
                f'move /Y "{new_exe_path}" "{current_exe}"',
                f'start "" "{current_exe}"',
                "(goto) 2>nul & del \"%~f0\"",
            ]
            bat_content = "\r\n".join(bat_lines) + "\r\n"
            # 使用 GBK 编码，确保中文 Windows 的 cmd.exe 能正常执行
            with open(bat_path, "w", encoding="gbk", errors="replace") as f:
                f.write(bat_content)

            self.root.after(0, lambda: self._launch_updater_and_exit(bat_path, new_version))

        except Exception as e:
            # 清理临时文件，避免留下损坏的文件
            for p in (new_exe_path, bat_path):
                if os.path.exists(p):
                    try:
                        os.remove(p)
                    except OSError:
                        pass
            msg = str(e)
            # 404 时给出明确指引
            if "404" in msg or "Not Found" in msg.lower():
                msg = (f"下载地址不存在 (404)，请检查：\n"
                       f"1. Release 标签名是否正确（如 v2.1）\n"
                       f"2. 上传的 EXE 文件名是否与 exe_url 一致\n"
                       f"3. 仓库是否为公开仓库\n\n"
                       f"当前 exe_url:\n{exe_url}")
            self.root.after(0, lambda: self._on_update_error(f"下载失败: {msg}"))

    def _launch_updater_and_exit(self, bat_path, new_version):
        """下载完成：确认后以隐藏窗口启动批处理脚本，然后退出主程序"""
        self.update_btn.config(state=tk.NORMAL, text="🔄 检查更新")
        if not messagebox.askyesno(
            "下载完成",
            f"新版本 {new_version} 下载完毕！\n\n"
            f"点击「是」程序将立即关闭并自动完成安装，安装后自动重新启动。\n"
            f"点击「否」取消本次安装（下载文件已保存，下次可重新触发）。"
        ):
            self.update_version_label.config(
                text=f"当前版本: {CURRENT_VERSION}  已取消安装"
            )
            return

        # 以隐藏控制台窗口的方式启动批处理，避免黑色命令行窗口弹出
        startupinfo = subprocess.STARTUPINFO()
        startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
        startupinfo.wShowWindow = subprocess.SW_HIDE
        subprocess.Popen(
            ["cmd.exe", "/C", bat_path],
            startupinfo=startupinfo,
            creationflags=subprocess.CREATE_NEW_PROCESS_GROUP
        )
        self.root.destroy()

def main():
    """主函数"""
    root = tk.Tk()
    app = ExcelMergerApp(root)
    root.mainloop()

if __name__ == "__main__":
    # 测试用的模拟数据创建函数
    def create_test_files():
        """创建测试用的Excel文件（仅用于演示）"""
        import pandas as pd
        
        # 测试数据1
        data1 = {
            '商品': ['商品A', '商品B', '商品C', '商品D'],
            '盒码': ['001', '002', '003', '004'],
            '订单量': [10, 20, 15, 5]
        }
        
        # 测试数据2
        data2 = {
            '商品': ['商品C', '商品D', '商品E', '商品F'],
            '盒码': ['003', '004', '005', '006'],
            '订单量': [5, 10, 8, 12]
        }
        
        df1 = pd.DataFrame(data1)
        df2 = pd.DataFrame(data2)
        
        df1.to_excel("测试文件1.xlsx", index=False)
        df2.to_excel("测试文件2.xlsx", index=False)
        print("测试文件已创建: 测试文件1.xlsx, 测试文件2.xlsx")
    
    # 运行主程序
    main()

